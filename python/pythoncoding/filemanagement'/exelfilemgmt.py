from openpyxl import Workbook, load_workbook
import os

def write_excel(filename):
    workbook=Workbook()
    sheet = workbook.active
    sheet.append(["name",'age'])
    sheet.append(["john",34])
    workbook.save(filename)

def read_excel(filename):
    workbook= load_workbook(filename)
    sheet=workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(f"Namw: {row[0]} , Age: {row[1]}")

def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print('file got deleted')
    else:
        print("file dost not exist")

filename="myfile.xlsx"
write_excel(filename)
print("Data read from csv file:")
read_excel(filename)
delete_excel(filename)
